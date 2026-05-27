"""
数据导入导出服务
支持 Excel (.xlsx) 和 CSV (.csv) 格式
"""
import csv
import io
import os
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

# 各记录类型的字段定义
RECORD_FIELDS = {
    "sleep": {
        "display_name": "睡眠记录",
        "fields": ["id", "start_time", "end_time", "sleep_type", "quality", "duration_minutes", "notes", "is_ongoing", "created_at"],
        "headers": ["ID", "开始时间", "结束时间", "睡眠类型", "质量(1-5)", "时长(分钟)", "备注", "进行中", "创建时间"],
        "date_field": "start_time",
    },
    "diaper": {
        "display_name": "排泄记录",
        "fields": ["id", "time", "diaper_type", "poop_color", "poop_consistency", "amount", "notes", "created_at"],
        "headers": ["ID", "时间", "类型", "便便颜色", "便便质地", "量", "备注", "创建时间"],
        "date_field": "time",
    },
    "cry": {
        "display_name": "哭声记录",
        "fields": ["id", "start_time", "end_time", "reason", "intensity", "soothing_method", "duration_minutes", "notes", "created_at"],
        "headers": ["ID", "开始时间", "结束时间", "原因", "强度(1-5)", "安抚方式", "时长(分钟)", "备注", "创建时间"],
        "date_field": "start_time",
    },
    "feeding": {
        "display_name": "喂养记录",
        "fields": ["id", "time", "feeding_type", "duration_minutes", "amount_ml", "breast_side", "solid_food", "water_amount_ml", "notes", "created_at"],
        "headers": ["ID", "时间", "喂养类型", "时长(分钟)", "奶量(ml)", "哺乳侧", "辅食", "水量(ml)", "备注", "创建时间"],
        "date_field": "time",
    },
    "growth": {
        "display_name": "生长发育记录",
        "fields": ["id", "record_date", "weight_kg", "height_cm", "head_circumference_cm", "temperature_c", "notes", "created_at"],
        "headers": ["ID", "记录日期", "体重(kg)", "身高(cm)", "头围(cm)", "体温(℃)", "备注", "创建时间"],
        "date_field": "record_date",
    },
    "reminder": {
        "display_name": "提醒记录",
        "fields": ["id", "title", "reminder_type", "reminder_date", "reminder_time", "repeat_type", "status", "notes", "created_at"],
        "headers": ["ID", "标题", "类型", "提醒日期", "提醒时间", "重复", "状态", "备注", "创建时间"],
        "date_field": "reminder_date",
    },
}


def export_to_csv(records: List[Dict], record_type: str) -> str:
    """导出记录为 CSV 字符串"""
    if record_type not in RECORD_FIELDS:
        raise ValueError(f"不支持的记录类型: {record_type}")

    config = RECORD_FIELDS[record_type]
    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    writer.writerow(config["headers"])

    # 写入数据
    for record in records:
        row = []
        for field in config["fields"]:
            value = record.get(field, "")
            if isinstance(value, bool):
                value = "是" if value else "否"
            elif value is None:
                value = ""
            row.append(value)
        writer.writerow(row)

    return output.getvalue()


def export_to_excel(records: List[Dict], record_type: str, output_path: Optional[str] = None) -> bytes:
    """导出记录为 Excel 字节流"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if record_type not in RECORD_FIELDS:
        raise ValueError(f"不支持的记录类型: {record_type}")

    config = RECORD_FIELDS[record_type]
    wb = Workbook()
    ws = wb.active
    ws.title = config["display_name"]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, header in enumerate(config["headers"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(config["fields"], 1):
            value = record.get(field, "")
            if isinstance(value, bool):
                value = "是" if value else "否"
            elif value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 自动调整列宽
    for col_idx, header in enumerate(config["headers"], 1):
        max_length = len(str(header))
        for row_idx in range(2, len(records) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 40)

    # 保存到字节流
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def import_from_csv(csv_content: str, record_type: str) -> List[Dict]:
    """从 CSV 导入记录"""
    if record_type not in RECORD_FIELDS:
        raise ValueError(f"不支持的记录类型: {record_type}")

    config = RECORD_FIELDS[record_type]
    reader = csv.DictReader(io.StringIO(csv_content))

    records = []
    for row in reader:
        record = {}
        for i, field in enumerate(config["fields"]):
            if field == "id":
                continue  # 跳过 ID，让系统自动生成
            header = config["headers"][i]
            value = row.get(header, "")
            if value == "":
                value = None
            record[field] = value
        records.append(record)

    return records


def import_from_excel(file_content: bytes, record_type: str) -> List[Dict]:
    """从 Excel 导入记录"""
    from openpyxl import load_workbook

    if record_type not in RECORD_FIELDS:
        raise ValueError(f"不支持的记录类型: {record_type}")

    config = RECORD_FIELDS[record_type]
    wb = load_workbook(io.BytesIO(file_content))
    ws = wb.active

    records = []
    for row_idx in range(2, ws.max_row + 1):
        record = {}
        for col_idx, field in enumerate(config["fields"], 1):
            if field == "id":
                continue
            value = ws.cell(row=row_idx, column=col_idx).value
            if value == "" or value is None:
                value = None
            record[field] = value
        records.append(record)

    return records


def get_supported_types() -> List[Dict]:
    """获取支持的导入导出类型"""
    return [
        {"key": k, "name": v["display_name"]}
        for k, v in RECORD_FIELDS.items()
    ]
